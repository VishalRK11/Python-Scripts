__author__ = 'Vishal'

import requests
import openpyxl as xl
from bs4 import BeautifulSoup as Soup
from fake_useragent import UserAgent

def generate_leader_board(url):

    user_agent = UserAgent()

    page = requests.get(url=url, headers={'user-agent': user_agent.chrome})

    soup = Soup(page.content, 'html.parser')

    wb = xl.Workbook()
    ws = wb.active
    ws.title = 'Leader board Batch 6'

    headings = [th.string for th in soup.find_all('th')]
    headings[0] = 'Batch Rank'
    group_names = ['sneha.vps888', 'shreyasndiddi', 'anirudhveeramalla11', 'naveen.yelisetti', 'kosuru.sasanka',
                   'samaikya888', 'sruppuluri1998']

    for idx, value in enumerate(headings):
        ws.cell(row=1, column=idx + 1, value=value.strip())

    tr = soup.find_all('tr')[174:226]
    row_value = 2

    for row in tr:
        td = row.find_all('td')
        text = td[1].text
        name = text[len(text)-text[::-1].find('_'):]
        if name.lower() in group_names:
            ws.cell(row=row_value, column=1, value=row_value-1)
            ws.cell(row=row_value, column=2, value=name.title())
            for col_value in range(3, 7):
                ws.cell(row=row_value, column=col_value, value=td[col_value-1].text)
            row_value += 1

    wb.save('Batch 6 Leader board.xlsx')

if __name__ == '__main__':
    generate_leader_board("https://s3-us-west-2.amazonaws.com/ppresults/ol2019/test_results.gz2018-07-23-19-48.html")
